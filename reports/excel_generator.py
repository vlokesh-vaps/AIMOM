"""Excel MoM generator using openpyxl.

Creates a single-sheet workbook aligned to the requested MoM tracker layout.
"""

import time
from pathlib import Path

from openpyxl import Workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter

from config.settings import COMPANY_THEME_COLOR
from utils.logger import get_logger

logger = get_logger(__name__)


class ExcelGenerator:
    """Generate a single-sheet action-oriented MoM workbook."""

    def __init__(self) -> None:
        self._output_dir = Path(__file__).resolve().parent / "excel"
        self._output_dir.mkdir(parents=True, exist_ok=True)

    def generate(self, summary_data: dict, filename: str) -> Path:
        start_time = time.time()
        output_path = self._output_dir / filename

        wb = Workbook()
        ws = wb.active
        ws.title = "MoM"
        ws.views.sheetView[0].showGridLines = True

        styles = self._make_styles()
        headers = ["Sl. No.", "Topic", "Detailed Action Item", "Owner", "Timeline"]
        ws.append(headers)
        self._style_header_row(ws, styles)

        rows = self._build_rows(summary_data)
        if not rows:
            rows = [["-", "No agenda items captured.", "", "", ""]]

        for row_index, row in enumerate(rows, start=2):
            ws.append(row)
            self._style_data_row(ws, row_index, styles)

        ws.freeze_panes = "A2"
        ws.auto_filter.ref = f"A1:E{max(ws.max_row, 2)}"
        self._set_widths(ws, {"A": 10, "B": 30, "C": 95, "D": 34, "E": 20})

        wb.save(str(output_path))

        duration = time.time() - start_time
        logger.info(
            "Excel generation complete. Path=%s, Rows=%d, Duration=%.2fs",
            output_path,
            ws.max_row,
            duration,
        )
        return output_path

    @staticmethod
    def _build_rows(summary_data: dict) -> list[list[str]]:
        discussion_points = summary_data.get("discussion_points", [])
        action_items = summary_data.get("action_items", [])

        actions_by_agenda: dict[str, list] = {}
        for item in action_items:
            agenda = ExcelGenerator._get(item, "agenda_item", "").strip() or "Off Agenda Discussion"
            actions_by_agenda.setdefault(agenda, []).append(item)

        rows: list[list[str]] = []
        used_action_ids: set[int] = set()

        for index, dp in enumerate(discussion_points, start=1):
            agenda = ExcelGenerator._get(dp, "agenda_item", "Off Agenda Discussion").strip() or "Off Agenda Discussion"
            matched_actions = actions_by_agenda.get(agenda, [])
            action_text = ExcelGenerator._build_action_text(dp, matched_actions)
            owners = ExcelGenerator._join_unique(ExcelGenerator._get(ai, "owner", "") for ai in matched_actions)
            timeline = ExcelGenerator._join_unique(ExcelGenerator._get(ai, "target_date", "") for ai in matched_actions)

            rows.append([
                index,
                agenda,
                action_text,
                owners or ExcelGenerator._get(dp, "assigned_to", "Not Specified"),
                timeline or ExcelGenerator._get(dp, "deadline", "Not Specified"),
            ])

            for ai in matched_actions:
                used_action_ids.add(id(ai))

        next_index = len(rows) + 1
        for item in action_items:
            if id(item) in used_action_ids:
                continue
            agenda = ExcelGenerator._get(item, "agenda_item", "").strip() or "Off Agenda Discussion"
            rows.append([
                next_index,
                agenda,
                ExcelGenerator._get(item, "task", ""),
                ExcelGenerator._get(item, "owner", ""),
                ExcelGenerator._get(item, "target_date", ""),
            ])
            next_index += 1

        return rows

    @staticmethod
    def _build_action_text(discussion_point, action_items: list) -> str:
        parts: list[str] = []

        point = ExcelGenerator._get(discussion_point, "point", "").strip()
        summary = ExcelGenerator._get(discussion_point, "detailed_summary", "").strip()
        decision = ExcelGenerator._get(discussion_point, "decision", "").strip()

        if point:
            parts.append(point)
        if summary and summary != point:
            parts.append(summary)
        if decision and decision != "No Decision Taken":
            parts.append(f"Decision: {decision}")

        tasks = [ExcelGenerator._get(item, "task", "").strip() for item in action_items]
        tasks = [task for task in tasks if task]
        if tasks:
            if parts:
                parts.append("")
            parts.extend(tasks)
        elif not parts:
            fallback_task = ExcelGenerator._get(discussion_point, "task", "").strip()
            if fallback_task and fallback_task != "No Action Item":
                parts.append(fallback_task)

        return "\n".join(parts).strip()

    @staticmethod
    def _make_styles() -> dict:
        theme_hex = COMPANY_THEME_COLOR.lstrip("#")
        if len(theme_hex) != 6:
            theme_hex = "1E3A8A"

        thin_border = Border(
            left=Side(style="thin", color="D9D9D9"),
            right=Side(style="thin", color="D9D9D9"),
            top=Side(style="thin", color="D9D9D9"),
            bottom=Side(style="thin", color="D9D9D9"),
        )
        return {
            "header_font": Font(name="Calibri", size=11, bold=True, color="FFFFFF"),
            "header_fill": PatternFill(start_color=theme_hex, end_color=theme_hex, fill_type="solid"),
            "data_font": Font(name="Calibri", size=10, color="000000"),
            "center": Alignment(horizontal="center", vertical="center", wrap_text=True),
            "left": Alignment(horizontal="left", vertical="top", wrap_text=True),
            "border": thin_border,
        }

    @staticmethod
    def _style_header_row(ws, styles: dict) -> None:
        for column in range(1, 6):
            cell = ws.cell(row=1, column=column)
            cell.font = styles["header_font"]
            cell.fill = styles["header_fill"]
            cell.alignment = styles["center"]
            cell.border = styles["border"]
        ws.row_dimensions[1].height = 24

    @staticmethod
    def _style_data_row(ws, row_index: int, styles: dict) -> None:
        for column in range(1, 6):
            cell = ws.cell(row=row_index, column=column)
            cell.font = styles["data_font"]
            cell.border = styles["border"]
            cell.alignment = styles["center"] if column == 1 else styles["left"]
        ws.row_dimensions[row_index].height = 48

    @staticmethod
    def _set_widths(ws, widths: dict[str, float]) -> None:
        for column, width in widths.items():
            ws.column_dimensions[column].width = width
        for col in ws.columns:
            col_letter = get_column_letter(col[0].column)
            if col_letter not in widths:
                ws.column_dimensions[col_letter].width = 18

    @staticmethod
    def _join_unique(values) -> str:
        seen: list[str] = []
        for value in values:
            text = str(value or "").strip()
            if not text or text in ("Not Specified", "-"):
                continue
            if text not in seen:
                seen.append(text)
        return "\n".join(seen)

    @staticmethod
    def _get(item, key: str, default: str = "") -> str:
        if isinstance(item, dict):
            return item.get(key, default)
        return getattr(item, key, default)
